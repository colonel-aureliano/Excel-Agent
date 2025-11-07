"""
Action Executor - Executes parsed actions on local Excel files
Provides a bridge between agent action strings and actual Excel file manipulation
"""

import re
from typing import Optional, Tuple, List
from .action_interpret import Action as ExcelAction


class ActionExecutor:
    """
    Executes actions on Excel files based on agent commands.
    Maintains the Excel Action instance and handles action parsing/execution.
    """
    
    def __init__(self, excel_file_path: str):
        """
        Initialize the action executor with an Excel file.
        
        Args:
            excel_file_path: Path to the Excel file to manipulate
        """
        self.excel_file_path = excel_file_path
        self.excel_action = ExcelAction(excel_file_path)
        self.last_operation_result = ""
        
    def parse_and_execute(self, action_string: str) -> Tuple[bool, str]:
        """
        Parse an action string from the agent and execute it.
        
        Args:
            action_string: The action string from the agent (e.g., "Select(0, 0, 4, 0)")
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        action_string = action_string.strip()
        
        # Handle Tell User actions
        if "Tell User" in action_string or "TellUser" in action_string:
            return True, "User interaction required"
            
        # Handle Terminate actions
        if "Terminate" in action_string:
            return True, "Task terminated"
        
        # Parse and execute Select actions with 4 parameters (range)
        select_match = re.search(r'Select\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', action_string, re.IGNORECASE)
        if select_match:
            col1_idx, row1_idx, col2_idx, row2_idx = map(int, select_match.groups())
            return self._execute_select(col1_idx, row1_idx, col2_idx, row2_idx)
        
        # Parse and execute Select actions with 2 parameters (single cell)
        select_single_match = re.search(r'Select\s*\(\s*(\d+)\s*,\s*(\d+)\s*\)', action_string, re.IGNORECASE)
        if select_single_match:
            col_idx, row_idx = map(int, select_single_match.groups())
            return self._execute_select(col_idx, row_idx, col_idx, row_idx)
        
        # Parse Select with Excel-style references (e.g., "Select A1:B5")
        select_excel_match = re.search(r'Select\s+([A-Z]+)(\d+)(?::([A-Z]+)(\d+))?', action_string, re.IGNORECASE)
        if select_excel_match:
            col1, row1, col2, row2 = select_excel_match.groups()
            col1_idx = self._column_to_index(col1)
            row1_idx = int(row1) - 1
            col2_idx = self._column_to_index(col2) if col2 else col1_idx
            row2_idx = int(row2) - 1 if row2 else row1_idx
            return self._execute_select(col1_idx, row1_idx, col2_idx, row2_idx)
        
        # Parse and execute SelectAndDrag actions
        drag_match = re.search(r'(?:Select\s*and\s*Drag|SelectAndDrag)\s*\(\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*\)', action_string, re.IGNORECASE)
        if drag_match:
            col1_idx, row1_idx, col2_idx, row2_idx = map(int, drag_match.groups())
            return self._execute_select_and_drag(col1_idx, row1_idx, col2_idx, row2_idx)
        
        # Parse and execute Set actions
        set_match = re.search(r'Set\s*\(\s*["\'](.+?)["\']\s*\)', action_string, re.IGNORECASE | re.DOTALL)
        if set_match:
            text = set_match.group(1)
            return self._execute_set(text)
        
        # Parse Set without quotes
        set_no_quote_match = re.search(r'Set\s*\(\s*([^)]+)\s*\)', action_string, re.IGNORECASE)
        if set_no_quote_match:
            text = set_no_quote_match.group(1).strip()
            return self._execute_set(text)
        
        # Parse and execute Format actions
        format_match = re.search(r'Format\s*\(\s*(.+?)\s*\)', action_string, re.IGNORECASE)
        if format_match:
            format_params = format_match.group(1)
            return self._execute_format(format_params)
        
        # Parse and execute Tool actions (copy, paste, delete, etc.)
        tool_match = re.search(r'(Copy|Paste|Delete|Cut|Bold|Italic|Underline)\s*\(\s*\)', action_string, re.IGNORECASE)
        if tool_match:
            tool_name = tool_match.group(1).lower()
            return self._execute_tool(tool_name)
        
        # Handle "Format Cells" and similar formatting dialog actions
        # These are no-ops since Select already applies highlighting
        if re.search(r'Format\s+Cells|Apply\s+Formatting|Background\s+Color', action_string, re.IGNORECASE):
            # The formatting was already applied during Select, so just return success
            if self.excel_action.selected_range:
                return True, "Formatting already applied during selection (cells are highlighted in yellow)"
            else:
                return False, "No cells selected for formatting"
        
        # Handle generic dialog/menu actions that don't need implementation
        if re.search(r'Copy/Paste/Delete/\.\.\.|Right\s*Click|Context\s+Menu', action_string, re.IGNORECASE):
            # These are interface actions that don't translate to file operations
            # If there's a selected range, consider it successful
            if self.excel_action.selected_range:
                return True, "Selection is already highlighted and formatted"
            else:
                return True, "Menu action acknowledged (no file changes needed)"
        
        # If we can't parse the action, log it and return False
        print(f"⚠️ Warning: Could not parse action format: {action_string}")
        return False, f"Could not parse action: {action_string}"
    
    def _column_to_index(self, col_letter: str) -> int:
        """Convert Excel column letter to 0-based index."""
        from openpyxl.utils import column_index_from_string
        return column_index_from_string(col_letter) - 1
    
    def _execute_select(self, col1_idx: int, row1_idx: int, col2_idx: int, row2_idx: int) -> Tuple[bool, str]:
        """
        Execute a Select action on the Excel file.
        
        Args:
            col1_idx: Starting column index (0-based from agent)
            row1_idx: Starting row index (0-based from agent)
            col2_idx: Ending column index (0-based from agent)
            row2_idx: Ending row index (0-based from agent)
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            # Convert 0-based indices to 1-based for action_interpret.py
            # action_interpret expects 1-based row and column numbers
            result = self.excel_action.select(col1_idx + 1, row1_idx + 1, col2_idx + 1, row2_idx + 1)
            self.last_operation_result = result
            
            # Apply highlighting to the selected range using openpyxl
            from openpyxl.styles import PatternFill
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for row in range(row1_idx, row2_idx + 1):
                for col in range(col1_idx, col2_idx + 1):
                    # openpyxl uses 1-based indexing
                    cell = self.excel_action.active_sheet.cell(row=row+1, column=col+1)
                    cell.fill = yellow_fill
            
            return True, result
        except Exception as e:
            error_msg = f"Error executing select: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg
    
    def _execute_select_and_drag(self, col1_idx: int, row1_idx: int, col2_idx: int, row2_idx: int) -> Tuple[bool, str]:
        """
        Execute a SelectAndDrag action (auto-fill).
        
        Args:
            col1_idx: Starting column index (0-based from agent)
            row1_idx: Starting row index (0-based from agent)
            col2_idx: Ending column index (0-based from agent)
            row2_idx: Ending row index (0-based from agent)
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            # Convert 0-based to 1-based for action_interpret.py
            result = self.excel_action.select_and_drag(col1_idx + 1, row1_idx + 1, col2_idx + 1, row2_idx + 1)
            self.last_operation_result = result
            return True, result
        except Exception as e:
            error_msg = f"Error executing select and drag: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            return False, error_msg
    
    def _execute_set(self, text: str) -> Tuple[bool, str]:
        """
        Execute a Set action (set cell value).
        
        Args:
            text: The text/value to set in the selected cells
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            # First select the input field, then set the value
            self.excel_action.select_input_field()
            result = self.excel_action.set_input(text)
            self.last_operation_result = result
            return True, result
        except Exception as e:
            error_msg = f"Error executing set: {str(e)}"
            print(error_msg)
            return False, error_msg
    
    def _execute_format(self, format_params: str) -> Tuple[bool, str]:
        """
        Execute a Format action (apply formatting to selected cells).
        
        Args:
            format_params: Format parameters (e.g., "bold", "color=#FF0000")
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            # Parse format parameters
            params = {}
            if "bold" in format_params.lower():
                params["bold"] = True
            if "italic" in format_params.lower():
                params["italic"] = True
            if "underline" in format_params.lower():
                params["underline"] = True
            
            # Color parsing
            color_match = re.search(r'color\s*=\s*["\']?([#\w]+)["\']?', format_params, re.IGNORECASE)
            if color_match:
                params["color"] = color_match.group(1)
            
            # Apply formatting using openpyxl
            if self.excel_action.selected_range:
                from openpyxl.styles import Font, PatternFill
                
                start_row, start_col, end_row, end_col = self.excel_action.selected_range
                
                for row in range(start_row, end_row + 1):
                    for col in range(start_col, end_col + 1):
                        cell = self.excel_action.active_sheet.cell(row=row+1, column=col+1)
                        
                        if params.get("bold"):
                            cell.font = Font(bold=True)
                        if params.get("italic"):
                            cell.font = Font(italic=True)
                        if params.get("underline"):
                            cell.font = Font(underline="single")
                        if params.get("color"):
                            color = params["color"].replace("#", "")
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                
                result = f"Applied formatting: {format_params}"
                self.last_operation_result = result
                return True, result
            else:
                return False, "No range selected for formatting"
                
        except Exception as e:
            error_msg = f"Error executing format: {str(e)}"
            print(error_msg)
            return False, error_msg
    
    def _execute_tool(self, tool_name: str) -> Tuple[bool, str]:
        """
        Execute a tool action (copy, paste, delete, etc.).
        
        Args:
            tool_name: Name of the tool (e.g., "copy", "paste", "delete")
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            tool_method_map = {
                "copy": "select_context_menu_tool",
                "paste": "select_context_menu_tool",
                "delete": "select_context_menu_tool",
                "bold": "select_top_menu_tool",
                "italic": "select_top_menu_tool",
                "underline": "select_top_menu_tool"
            }
            
            method_name = tool_method_map.get(tool_name)
            if method_name:
                method = getattr(self.excel_action, method_name)
                result = method(tool_name)
                self.last_operation_result = result
                return True, result
            else:
                return False, f"Unknown tool: {tool_name}"
                
        except Exception as e:
            error_msg = f"Error executing tool {tool_name}: {str(e)}"
            print(error_msg)
            return False, error_msg
    
    def save(self, output_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        Save the Excel file.
        
        Args:
            output_path: Optional output path (defaults to original file path)
            
        Returns:
            Tuple of (success: bool, result_message: str)
        """
        try:
            result = self.excel_action.save(output_path)
            return True, result
        except Exception as e:
            error_msg = f"Error saving file: {str(e)}"
            print(error_msg)
            return False, error_msg
    
    def get_df(self):
        """Get the current DataFrame."""
        return self.excel_action.df
    
    def get_workbook(self):
        """Get the current openpyxl workbook."""
        return self.excel_action.workbook
    
    def get_sheet_summary(self, num_rows: int = 10) -> str:
        """
        Get a summary of the current sheet state.
        
        Args:
            num_rows: Number of rows to include in summary
            
        Returns:
            String summary of the sheet
        """
        df = self.excel_action.df
        summary = f"Excel Sheet Summary:\n"
        summary += f"Shape: {df.shape[0]} rows x {df.shape[1]} columns\n"
        summary += f"Columns: {list(df.columns)}\n\n"
        summary += f"First {min(num_rows, len(df))} rows:\n"
        summary += df.head(num_rows).to_string()
        return summary

