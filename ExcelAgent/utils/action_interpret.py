import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
import re
import os
from typing import Union, List, Tuple, Optional, Dict, Any
import numpy as np

# Interpretation of actions on Excel Spreadsheets using pandas and openpyxl

class Action:
    def __init__(self, file_path: str):
        """
        Initialize the Excel Agent with a specified Excel file.
        
        Args:
            file_path: Path to the Excel file
        """
        self.file_path = file_path
        # Load with pandas for data manipulation, explicitly specifying the engine
        self.df = pd.read_excel(file_path, engine='openpyxl')  # or 'xlrd' for .xls files
        # Load with openpyxl for formatting and other Excel-specific operations
        self.workbook = openpyxl.load_workbook(file_path)
        self.active_sheet = self.workbook.active
        self.selected_range = None
        self.clipboard = None
        self.input_field_content = ""
        self.input_field_selected = False
        
    def save(self, output_path: Optional[str] = None):
        try:
            save_path = output_path if output_path else self.file_path
            
            # Debug information
            print(f"DataFrame shape before saving: {self.df.shape}")
            print(f"DataFrame has data: {not self.df.empty}")
            
            # Try both methods for redundancy
            try:
                # Save with pandas
                self.df.to_excel(save_path, index=False, engine='openpyxl')
                print("Saved with pandas successfully")
            except Exception as e:
                print(f"Error saving with pandas: {e}")
            
            try:
                # Also save with openpyxl (with data sync)
                for r in range(len(self.df)):
                    for c in range(len(self.df.columns)):
                        cell = self.active_sheet.cell(row=r+1, column=c+1)
                        cell.value = self.df.iloc[r, c]
                
                self.workbook.save(save_path)
                print("Saved with openpyxl successfully")
            except Exception as e:
                print(f"Error saving with openpyxl: {e}")
            
            return f"File saved to {save_path}"
        except Exception as e:
            print(f"Unexpected error during save: {e}")
            raise
    
    def _convert_excel_ref(self, col: Union[str, int], row: Union[str, int]) -> Tuple[int, int]:
        """
        Convert Excel-style references to zero-indexed row and column indices.
        
        Args:
            col: Column reference (either as letter 'A' or number 1)
            row: Row reference (1-indexed as in Excel)
            
        Returns:
            Tuple of (row_idx, col_idx) zero-indexed for pandas
        """
        # Convert column to index if it's a letter
        if isinstance(col, str) and col.isalpha():
            col_idx = column_index_from_string(col) - 1
        else:
            # Convert to 0-indexed
            col_idx = int(col) - 1
        
        # Convert row to 0-indexed
        row_idx = int(row) - 1
        
        return row_idx, col_idx
    
    def _get_cell_address(self, row_idx: int, col_idx: int) -> str:
        """
        Convert zero-indexed row and column indices to Excel-style cell reference.
        
        Args:
            row_idx: Zero-indexed row
            col_idx: Zero-indexed column
            
        Returns:
            Excel-style cell reference (e.g., 'A1')
        """
        col_letter = get_column_letter(col_idx + 1)
        return f"{col_letter}{row_idx + 1}"
    
    def select(self, col1: Union[str, int], row1: Union[str, int], 
               col2: Optional[Union[str, int]] = None, row2: Optional[Union[str, int]] = None,
               condition: Optional[str] = None) -> str:
        """
        Select a range of cells, a column, or a row.
        
        Args:
            col1: Starting column
            row1: Starting row
            col2: Ending column (optional, defaults to col1)
            row2: Ending row (optional, defaults to row1)
            condition: Conditional expression to filter selection (optional)
            
        Returns:
            Description of the selection
        """
        # If col2 or row2 not specified, use col1/row1
        col2 = col1 if col2 is None else col2
        row2 = row1 if row2 is None else row2
        
        # Convert to indices
        start_row, start_col = self._convert_excel_ref(col1, row1)
        end_row, end_col = self._convert_excel_ref(col2, row2)
        
        # Store the selection
        self.selected_range = (start_row, start_col, end_row, end_col)
        
        # Check if it's a whole column selection
        if (start_row == 0 and end_row >= len(self.df) - 1):
            selection_type = f"Column {get_column_letter(start_col + 1)}"
            if start_col != end_col:
                selection_type = f"Columns {get_column_letter(start_col + 1)} to {get_column_letter(end_col + 1)}"
        # Check if it's a whole row selection
        elif (start_col == 0 and end_col >= len(self.df.columns) - 1):
            selection_type = f"Row {start_row + 1}"
            if start_row != end_row:
                selection_type = f"Rows {start_row + 1} to {end_row + 1}"
        # Otherwise it's a cell range
        else:
            start_address = self._get_cell_address(start_row, start_col)
            end_address = self._get_cell_address(end_row, end_col)
            selection_type = f"Range {start_address}:{end_address}"
        
        # Apply condition if specified
        if condition:
            # Here we would implement condition filtering
            # This is a simplified version
            return f"Selected {selection_type} with condition: {condition}"
        
        return f"Selected {selection_type}"
    
    def select_and_drag(self, col1: Union[str, int], row1: Union[str, int],
                        col2: Union[str, int], row2: Union[str, int],
                        condition: Optional[str] = None) -> str:
        """
        Select a cell's fill handle and drag to fill a range.
        
        Args:
            col1: Starting column (source cell)
            row1: Starting row (source cell)
            col2: Ending column (target range)
            row2: Ending row (target range)
            condition: Conditional expression for fill (optional)
            
        Returns:
            Description of the fill operation
        """
        # Convert to indices
        source_row, source_col = self._convert_excel_ref(col1, row1)
        target_row, target_col = self._convert_excel_ref(col2, row2)
        
        # Get the source cell value
        source_value = self.df.iloc[source_row, source_col]
        source_address = self._get_cell_address(source_row, source_col)

        is_numeric = isinstance(source_value, (int, float, np.number))
        
        # Determine fill direction and pattern
        if source_row == target_row:  # Horizontal fill
            # Check if source is a number and implement a sequence
            if is_numeric:
                for c in range(source_col + 1, target_col + 1):
                    increment = c - source_col
                    self.df.iloc[source_row, c] = source_value + increment
            else:  # Copy the same value
                for c in range(source_col + 1, target_col + 1):
                    self.df.iloc[source_row, c] = source_value
        
        elif source_col == target_col:  # Vertical fill
            # Check if source is a number and implement a sequence
            if is_numeric:
                for r in range(source_row + 1, target_row + 1):
                    increment = r - source_row
                    self.df.iloc[r, source_col] = source_value + increment
            else:  # Copy the same value
                for r in range(source_row + 1, target_row + 1):
                    self.df.iloc[r, source_col] = source_value
        
        elif source_col == target_col:  # Vertical fill
            # Check if source is a number and implement a sequence
            if isinstance(source_value, (int, float)):
                for r in range(source_row + 1, target_row + 1):
                    increment = r - source_row
                    self.df.iloc[r, source_col] = source_value + increment
            else:  # Copy the same value
                for r in range(source_row + 1, target_row + 1):
                    self.df.iloc[r, source_col] = source_value
        
        else:  # Both horizontal and vertical fill
            for r in range(source_row, target_row + 1):
                for c in range(source_col, target_col + 1):
                    if r == source_row and c == source_col:
                        continue  # Skip the source cell
                    self.df.iloc[r, c] = source_value
        
        target_address = self._get_cell_address(target_row, target_col)
        
        # Apply condition if specified
        if condition:
            # Here we would implement conditional filling
            return f"Filled from {source_address} to {target_address} with condition: {condition}"
        
        return f"Filled from {source_address} to {target_address}"
    
    def select_input_field(self) -> str:
        """
        Select the input field (formula bar).
        
        Returns:
            Confirmation message
        """
        self.input_field_selected = True
        
        # If there's a cell selection, populate input field with its value
        if self.selected_range:
            start_row, start_col, _, _ = self.selected_range
            if start_row <= len(self.df) - 1 and start_col <= len(self.df.columns) - 1:
                self.input_field_content = str(self.df.iloc[start_row, start_col])
        
        return "Input field selected"
    
    def set_input(self, text: str) -> str:
        """
        Type something into the input field.
        
        Args:
            text: Text to enter into the input field
            
        Returns:
            Confirmation message
        """
        if not self.input_field_selected:
            return "Error: Input field is not selected"
        
        self.input_field_content = text
        
        # If there's a cell selection, update its value
        if self.selected_range:
            start_row, start_col, end_row, end_col = self.selected_range
            
            # Determine if it's a formula
            is_formula = text.startswith("=")
            
            # Handle simple formulas (basic implementation)
            if is_formula:
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        if r < len(self.df) and c < len(self.df.columns):
                            self.df.iloc[r, c] = text
                return f"Formula '{text}' entered in selected range"
            
            # For regular text/values
            try:
                # Try to convert to number if possible
                numeric_value = pd.to_numeric(text)
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        if r < len(self.df) and c < len(self.df.columns):
                            self.df.iloc[r, c] = numeric_value
            except:
                # Otherwise treat as text
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        if r < len(self.df) and c < len(self.df.columns):
                            self.df.iloc[r, c] = text
        
        start_address = self._get_cell_address(start_row, start_col)
        if start_row == end_row and start_col == end_col:
            return f"Set value '{text}' in cell {start_address}"
        else:
            end_address = self._get_cell_address(end_row, end_col)
            return f"Set value '{text}' in range {start_address}:{end_address}"
    
    def select_top_menu_tool(self, tool_name: str, parameters: Optional[Dict[str, Any]] = None) -> str:
        """
        Select a tool from the top menu.
        
        Args:
            tool_name: Name of the tool to select
            parameters: Optional parameters for the tool
            
        Returns:
            Description of the action taken
        """
        if not self.selected_range:
            return "Error: No range selected for formatting"
        
        start_row, start_col, end_row, end_col = self.selected_range
        
        # Handle different top menu tools
        if tool_name.lower() == "bold":
            # Apply bold formatting in openpyxl
            for r in range(start_row + 1, end_row + 2):  # +1 for 1-indexing in openpyxl
                for c in range(start_col + 1, end_col + 2):
                    cell = self.active_sheet.cell(row=r, column=c)
                    cell.font = openpyxl.styles.Font(bold=True)
            return "Applied bold formatting to selected range"
        
        elif tool_name.lower() == "italic":
            # Apply italic formatting
            for r in range(start_row + 1, end_row + 2):
                for c in range(start_col + 1, end_col + 2):
                    cell = self.active_sheet.cell(row=r, column=c)
                    cell.font = openpyxl.styles.Font(italic=True)
            return "Applied italic formatting to selected range"
        
        elif tool_name.lower() == "conditional_formatting":
            if not parameters:
                return "Error: Conditional formatting requires parameters"
            
            # Basic implementation of conditional formatting
            rule_type = parameters.get("rule_type", "")
            threshold = parameters.get("threshold", 0)
            
            if rule_type == "greater_than":
                # Use pandas for demonstration
                for r in range(start_row, end_row + 1):
                    for c in range(start_col, end_col + 1):
                        if r < len(self.df) and c < len(self.df.columns):
                            try:
                                if float(self.df.iloc[r, c]) > threshold:
                                    # Mark in openpyxl
                                    cell = self.active_sheet.cell(row=r+1, column=c+1)
                                    cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                            except:
                                pass  # Skip non-numeric cells
                
                return f"Applied conditional formatting for values > {threshold}"
            
            return f"Applied {rule_type} conditional formatting"
        
        elif tool_name.lower() == "number_format":
            '''
            Common format codes include:
            "0.00" (fixed decimal places)
            "$#,##0.00" (currency)
            "0%" (percentage)
            "mm/dd/yyyy" (date)
            "h:mm AM/PM" (time)
            '''
            format_code = parameters.get("format", "General")
            
            # Apply number formatting
            for r in range(start_row + 1, end_row + 2):
                for c in range(start_col + 1, end_col + 2):
                    cell = self.active_sheet.cell(row=r, column=c)
                    cell.number_format = format_code
            
            return f"Applied number format '{format_code}' to selected range"
        
        else:
            return f"Selected tool: {tool_name} (not implemented in this demo)"
    
    def select_context_menu_tool(self, tool_name: str) -> str:
        """
        Select a tool from the right-click context menu.
        
        Args:
            tool_name: Name of the tool to select
            
        Returns:
            Description of the action taken
        """
        if not self.selected_range:
            return "Error: No selection for context menu action"
        
        start_row, start_col, end_row, end_col = self.selected_range
        
        if tool_name.lower() == "copy":
            # Create a copy of the selected data
            self.clipboard = []
            for r in range(start_row, end_row + 1):
                row_data = []
                for c in range(start_col, end_col + 1):
                    if r < len(self.df) and c < len(self.df.columns):
                        row_data.append(self.df.iloc[r, c])
                self.clipboard.append(row_data)
            
            return "Copied selection to clipboard"
        
        elif tool_name.lower() == "paste":
            if not self.clipboard:
                return "Error: Nothing to paste"
            
            # Paste from the starting position of the selection
            for r_offset, row_data in enumerate(self.clipboard):
                for c_offset, value in enumerate(row_data):
                    r = start_row + r_offset
                    c = start_col + c_offset
                    if r < len(self.df) and c < len(self.df.columns):
                        self.df.iloc[r, c] = value
            
            return "Pasted data from clipboard"
        
        elif tool_name.lower() == "delete":
            # Delete contents of selected cells
            for r in range(start_row, end_row + 1):
                for c in range(start_col, end_col + 1):
                    if r < len(self.df) and c < len(self.df.columns):
                        self.df.iloc[r, c] = None
            
            return "Deleted contents of selected cells"
        
        elif tool_name.lower() == "clear_formatting":
            # Clear formatting from selected cells
            for r in range(start_row + 1, end_row + 2):
                for c in range(start_col + 1, end_col + 2):
                    cell = self.active_sheet.cell(row=r, column=c)
                    cell.font = openpyxl.styles.Font()
                    cell.fill = openpyxl.styles.PatternFill()
                    cell.number_format = "General"
            
            return "Cleared formatting from selected cells"
        
        else:
            return f"Selected context menu tool: {tool_name} (not implemented in this demo)"
    
    def tell_user(self, message: str) -> str:
        """
        Send a message to the user.
        
        Args:
            message: Message to send to the user
            
        Returns:
            The message
        """
        return message


# Example usage
if __name__ == "__main__":
    agent = Action("Excel-Agent/example.xlsx")

    if agent.df.empty:
        # Create a basic structure with default column names
        agent.df = pd.DataFrame(columns=['A', 'B', 'C', 'D'])
        # Resize to have enough rows
        for i in range(10):  # Create 10 empty rows
            agent.df.loc[i] = [None] * len(agent.df.columns)
    
    # Select cell A1
    print(agent.select("A", 1))
    
    # Set a value
    print(agent.select_input_field())
    print(agent.set_input("100"))
    
    # Select and drag to create a sequence
    print(agent.select("A", 1))
    print(agent.select_and_drag("A", 1, "A", 5))
    
    # Use a top menu tool
    print(agent.select("A", 1, "A", 5))
    print(agent.select_top_menu_tool("bold"))
    
    # Use a context menu tool
    print(agent.select("A", 1, "A", 5))
    print(agent.select_context_menu_tool("copy"))
    print(agent.select("B", 1))
    print(agent.select_context_menu_tool("paste"))
    
    # Tell the user something
    print(agent.tell_user("I've created a sequence of numbers in column A and copied it to column B"))
    
    # Save the workbook
    print(agent.save("result.xlsx"))